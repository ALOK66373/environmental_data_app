from flask import Flask, request, render_template, redirect, send_file, jsonify
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from io import BytesIO
import shutil

EXCEL_FILE = os.getenv("EXCEL_FILE", "environmental_data.xlsx")

app = Flask(__name__)
EXCEL_FILE = "environmental_data.xlsx"
BACKUP_FOLDER = "backups"
os.makedirs(BACKUP_FOLDER, exist_ok=True)

def is_file_locked(filepath):
    if not os.path.exists(filepath):
        return False
    try:
        os.rename(filepath, filepath)
        return False
    except OSError:
        return True

def is_excel_file_valid(filepath):
    try:
        pd.read_excel(filepath)
        return True
    except Exception:
        return False

def backup_excel_file():
    if os.path.exists(EXCEL_FILE):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(BACKUP_FOLDER, f"environmental_data_backup_{timestamp}.xlsx")
        shutil.copy2(EXCEL_FILE, backup_path)

def init_excel_file():
    print("Initializing new Excel file...")
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        pd.DataFrame(columns=["Date", "Fertile", "PH", "H2 - Ppb", "ORP -Mv", "Res", "us/cm", "TDS -ppm", "Salt %", "Salt sg"]).to_excel(writer, sheet_name="Water Test", index=False)
        pd.DataFrame(columns=["Date", "Moisture", "Light", "Ph"]).to_excel(writer, sheet_name="Soil Test", index=False)
        pd.DataFrame(columns=["Date", "Temperature", "Humidity", "Air pollution level", "PM2.5 (AQI)", "HCHO", "TVOC"]).to_excel(writer, sheet_name="Air Test", index=False)

def sort_and_save_excel(sheet):
    df = pd.read_excel(EXCEL_FILE, sheet_name=sheet)
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.dropna(subset=['Date']).sort_values('Date')
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet, index=False)
        worksheet = writer.sheets[sheet]
        for i, column in enumerate(df.columns):
            max_len = max(df[column].astype(str).map(len).max(), len(column)) + 2
            worksheet.column_dimensions[get_column_letter(i+1)].width = max(12, max_len)
    return df

@app.route('/')
def form():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    if is_file_locked(EXCEL_FILE):
        return "Excel file is currently open/locked. Please close it before submitting data.", 409

    data = request.form.to_dict()
    sheet = data.pop("sheet")
    try:
        input_date = pd.to_datetime(request.form.get("Date")).date()
        data["Date"] = input_date
    except Exception:
        return "Invalid or missing date format.", 400

    for key in data:
        if key == "Date":
            continue
        value = data[key]
        if value == '':
            data[key] = None
        else:
            try:
                data[key] = float(value)
            except ValueError:
                return f"Invalid input for {key}. Must be a number.", 400

    if not os.path.exists(EXCEL_FILE) or not is_excel_file_valid(EXCEL_FILE):
        init_excel_file()

    backup_excel_file()

    wb = load_workbook(EXCEL_FILE)
    if sheet not in wb.sheetnames:
        return f"Sheet '{sheet}' not found in workbook.", 400

    ws = wb[sheet]
    headers = [cell.value for cell in ws[1]]
    date_index = headers.index("Date")

    for row in ws.iter_rows(min_row=2):
        cell_value = row[date_index].value
        if isinstance(cell_value, datetime):
            cell_value = cell_value.date()
        if cell_value == input_date:
            ws.delete_rows(row[0].row)
            break

    row_data = [data.get(h, None) for h in headers]
    ws.append(row_data)
    wb.save(EXCEL_FILE)

    sort_and_save_excel(sheet)
    print(f"✅ Data saved and sorted for sheet: {sheet}, date: {input_date}")
    return redirect('/')

@app.route('/view/<sheet_name>')
def view_sheet(sheet_name):
    try:
        if not os.path.exists(EXCEL_FILE) or not is_excel_file_valid(EXCEL_FILE):
            return f"Excel file missing or corrupted. Please restart the server.", 500
        df = sort_and_save_excel(sheet_name)
        return render_template('view_sheet.html', sheet_name=sheet_name, columns=df.columns, rows=df.to_dict(orient='records'))
    except Exception as e:
        return f"Error: {e}", 500

@app.route('/edit', methods=['POST'])
def edit_sheet():
    try:
        if is_file_locked(EXCEL_FILE):
            return jsonify({"error": "Excel file is currently locked. Please try again."}), 409

        data = request.get_json()
        sheet = data['sheet']
        columns = data['columns']
        rows = data['data']

        df = pd.DataFrame(rows, columns=columns)
        for col in df.columns:
            if col != 'Date':
                try:
                    df[col] = pd.to_numeric(df[col], errors='ignore')
                except:
                    pass

        backup_excel_file()

        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)

        sort_and_save_excel(sheet)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/download/<sheet_name>')
def download_sheet(sheet_name):
    try:
        if not os.path.exists(EXCEL_FILE) or not is_excel_file_valid(EXCEL_FILE):
            return f"File {EXCEL_FILE} not found or corrupted.", 404
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        output.seek(0)
        filename = f"{sheet_name.replace(' ', '_')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return f"Error: {e}", 500
    
    
@app.route('/download-all')
def download_all_sheets():
    try:
        if not os.path.exists(EXCEL_FILE):
            return f"File {EXCEL_FILE} not found.", 404

        # Load the entire workbook using pandas
        xl = pd.ExcelFile(EXCEL_FILE)
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name in xl.sheet_names:
                df = xl.parse(sheet_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="environmental_data_full.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return f"Error: {e}", 500
    

@app.route('/plot')
def plot_page():
    return render_template('plots.html')


@app.route('/api/plot/<sheet>')
def plot_data(sheet):
    print(f"[DEBUG] Plot data requested for sheet: {sheet}")
    print(f"[DEBUG] Excel file path: {os.path.abspath(EXCEL_FILE)}")
    
    if not os.path.exists(EXCEL_FILE) or not is_excel_file_valid(EXCEL_FILE):
        print("[ERROR] Excel file not found or invalid.")
        return jsonify({"error": "Excel file not found or corrupted."}), 404
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet)
        print(f"[DEBUG] Rows read from sheet '{sheet}': {len(df)}")
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date']).sort_values('Date')
        df = df.fillna('')
        df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
        return jsonify(df.to_dict(orient='records'))
    except Exception as e:
        print(f"[ERROR] Exception while reading sheet '{sheet}': {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
